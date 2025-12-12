from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.db.models import Count, Q, F, Max, Prefetch
from django.db import transaction, models
from django.http import HttpResponse
import csv
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from apps.modulo_3.cursos.models import Curso, Comision, PoloCreativo, Material, ComisionDocente
from apps.modulo_2.inscripciones.models import Inscripcion
from apps.modulo_1.roles.models import Estudiante, Docente, Rol, UsuarioRol
from apps.modulo_1.usuario.models import Persona, Usuario
from apps.modulo_4.asistencia.models import Asistencia
from apps.modulo_3.cursos.forms import MaterialForm
from datetime import date


def es_admin(user):
    """Verifica si el usuario es administrador, mesa de entrada o docente con cursos asignados"""
    if not user.is_authenticated:
        return False
    
    # Verificar si es staff o superuser (administrador Django)
    if user.is_staff or user.is_superuser:
        return True
    
    # Verificar si tiene rol de Administrador o Mesa de Entrada
    try:
        usuario = Usuario.objects.get(persona__dni=user.username)
        roles = UsuarioRol.objects.filter(usuario_id=usuario).values_list('rol_id__nombre', flat=True)
        if 'Administrador' in roles or 'Mesa de Entrada' in roles:
            return True
        
        # Verificar si es docente con comisiones asignadas
        from apps.modulo_3.cursos.models import ComisionDocente
        if Docente.objects.filter(id_persona=usuario.persona).exists():
            tiene_comisiones = ComisionDocente.objects.filter(fk_id_docente=usuario).exists()
            if tiene_comisiones:
                return True
    except Usuario.DoesNotExist:
        pass
    
    return False


def es_admin_completo(user):
    """Verifica si el usuario es administrador completo (puede crear usuarios)"""
    if not user.is_authenticated:
        return False
    
    # Verificar si es staff o superuser (administrador Django)
    if user.is_staff or user.is_superuser:
        return True
    
    # Verificar si tiene rol de Administrador (no Mesa de Entrada)
    try:
        usuario = Usuario.objects.get(persona__dni=user.username)
        roles = UsuarioRol.objects.filter(usuario_id=usuario).values_list('rol_id__nombre', flat=True)
        return 'Administrador' in roles
    except Usuario.DoesNotExist:
        return False


@login_required
@user_passes_test(es_admin)
def panel_cursos(request):
    """Panel de gestión de cursos"""
    cursos = Curso.objects.all().annotate(
        total_comisiones=Count('comision')
    ).order_by('orden', 'id_curso')
    
    context = {
        'cursos': cursos,
    }
    return render(request, 'administracion/panel_cursos.html', context)


@login_required
@user_passes_test(es_admin)
def crear_curso(request):
    """Crear un nuevo curso"""
    if request.method == 'POST':
        try:
            nombre = request.POST.get('nombre')
            descripcion = request.POST.get('descripcion', '')
            edad_minima = request.POST.get('edad_minima')
            edad_maxima = request.POST.get('edad_maxima')
            requisitos = request.POST.get('requisitos', '')
            contenido_multimedia = request.POST.get('contenido_multimedia', '')
            
            orden = request.POST.get('orden', '0')
            # Si no se especifica orden, usar el máximo + 1
            if not orden or orden == '':
                max_orden = Curso.objects.aggregate(models.Max('orden'))['orden__max'] or 0
                orden = max_orden + 1
            
            Curso.objects.create(
                nombre=nombre,
                descripcion=descripcion,
                edad_minima=int(edad_minima) if edad_minima else None,
                edad_maxima=int(edad_maxima) if edad_maxima else None,
                requisitos=requisitos,
                contenido_multimedia=contenido_multimedia,
                estado='Abierto',
                orden=int(orden)
            )
            
            messages.success(request, f'✅ Curso "{nombre}" creado exitosamente.')
            return redirect('administracion:panel_cursos')
            
        except Exception as e:
            messages.error(request, f'❌ Error al crear curso: {str(e)}')
    
    return render(request, 'administracion/crear_curso.html')


@login_required
@user_passes_test(es_admin)
def editar_curso(request, curso_id):
    """Editar un curso existente"""
    curso = get_object_or_404(Curso, id_curso=curso_id)
    
    if request.method == 'POST':
        try:
            curso.nombre = request.POST.get('nombre')
            curso.descripcion = request.POST.get('descripcion', '')
            edad_minima = request.POST.get('edad_minima')
            curso.edad_minima = int(edad_minima) if edad_minima else None
            edad_maxima = request.POST.get('edad_maxima')
            curso.edad_maxima = int(edad_maxima) if edad_maxima else None
            curso.requisitos = request.POST.get('requisitos', '')
            curso.contenido_multimedia = request.POST.get('contenido_multimedia', '')
            curso.estado = request.POST.get('estado', 'Abierto')
            orden = request.POST.get('orden', '0')
            curso.orden = int(orden) if orden else 0
            curso.save()
            
            messages.success(request, f'✅ Curso "{curso.nombre}" actualizado exitosamente.')
            return redirect('administracion:panel_cursos')
            
        except Exception as e:
            messages.error(request, f'❌ Error al actualizar curso: {str(e)}')
    
    context = {'curso': curso}
    return render(request, 'administracion/editar_curso.html', context)


@login_required
@user_passes_test(es_admin)
def panel_comisiones(request):
    """Panel de gestión de comisiones"""
    from apps.modulo_3.cursos.models import ComisionDocente
    
    comisiones = Comision.objects.all().select_related('fk_id_curso', 'fk_id_polo').prefetch_related('docentes').order_by('-id_comision')
    
    # Obtener docentes asignados para cada comisión
    comisiones_con_docentes = []
    for comision in comisiones:
        docentes_asignados = ComisionDocente.objects.filter(fk_id_comision=comision).select_related('fk_id_docente__persona')
        comisiones_con_docentes.append({
            'comision': comision,
            'docentes': docentes_asignados,
        })
    
    # Obtener docentes disponibles para asignar
    docentes_disponibles = []
    try:
        docentes = Docente.objects.all().select_related('id_persona')
        for docente in docentes:
            usuario = Usuario.objects.filter(persona=docente.id_persona).first()
            if usuario:
                docentes_disponibles.append({
                    'id': usuario.id,
                    'nombre_completo': docente.id_persona.nombre_completo,
                    'dni': docente.id_persona.dni,
                    'especialidad': docente.especialidad,
                })
    except Exception as e:
        pass
    
    context = {
        'comisiones_con_docentes': comisiones_con_docentes,
        'docentes_disponibles': docentes_disponibles,
    }
    return render(request, 'administracion/panel_comisiones.html', context)


@login_required
@user_passes_test(es_admin)
def crear_comision(request):
    """Crear una nueva comisión"""
    from apps.modulo_3.cursos.models import ComisionDocente
    
    if request.method == 'POST':
        try:
            curso_id = request.POST.get('curso')
            polo_id = request.POST.get('polo')
            docente_id = request.POST.get('docente')  # ID del usuario docente
            
            curso = Curso.objects.get(id_curso=curso_id)
            polo = PoloCreativo.objects.get(id_polo=polo_id) if polo_id else None
            
            # Crear la comisión
            comision = Comision.objects.create(
                fk_id_curso=curso,
                fk_id_polo=polo,
                modalidad=request.POST.get('modalidad', 'Presencial'),
                dias_horarios=request.POST.get('dias_horarios', ''),
                lugar=request.POST.get('lugar', ''),
                fecha_inicio=request.POST.get('fecha_inicio') or None,
                fecha_fin=request.POST.get('fecha_fin') or None,
                cupo_maximo=int(request.POST.get('cupo_maximo', 20)),
                estado='Abierta'
            )
            
            # Asignar docente si se seleccionó uno
            if docente_id:
                try:
                    docente_usuario = Usuario.objects.get(id=docente_id)
                    # Verificar que el usuario sea realmente un docente
                    if Docente.objects.filter(id_persona=docente_usuario.persona).exists():
                        ComisionDocente.objects.get_or_create(
                            fk_id_comision=comision,
                            fk_id_docente=docente_usuario
                        )
                        messages.success(request, f'✅ Comisión creada exitosamente con docente asignado.')
                    else:
                        messages.warning(request, f'✅ Comisión creada exitosamente. ⚠️ El usuario seleccionado no es un docente válido.')
                except Usuario.DoesNotExist:
                    messages.warning(request, f'✅ Comisión creada exitosamente. ⚠️ No se pudo asignar el docente.')
            else:
                messages.success(request, f'✅ Comisión creada exitosamente. Puedes asignar un docente más tarde.')
            
            return redirect('administracion:panel_comisiones')
            
        except Exception as e:
            messages.error(request, f'❌ Error al crear comisión: {str(e)}')
    
    cursos = Curso.objects.filter(estado='Abierto')
    polos = PoloCreativo.objects.filter(activo=True)
    
    # Obtener docentes disponibles (usuarios que tienen el rol de Docente)
    docentes_disponibles = []
    try:
        # Obtener todas las personas que son docentes
        docentes = Docente.objects.all().select_related('id_persona')
        for docente in docentes:
            # Obtener el usuario asociado a la persona
            usuario = Usuario.objects.filter(persona=docente.id_persona).first()
            if usuario:
                docentes_disponibles.append({
                    'id': usuario.id,
                    'nombre_completo': docente.id_persona.nombre_completo,
                    'dni': docente.id_persona.dni,
                    'especialidad': docente.especialidad,
                })
    except Exception as e:
        pass
    
    context = {
        'cursos': cursos,
        'polos': polos,
        'docentes_disponibles': docentes_disponibles,
    }
    return render(request, 'administracion/crear_comision.html', context)


@login_required
@user_passes_test(es_admin)
def asignar_docente_comision(request, comision_id):
    """Asignar un docente a una comisión"""
    # DESHABILITADO TEMPORALMENTE
    messages.warning(request, 'La asignación de docentes no está habilitada por el momento.')
    return redirect('administracion:panel_comisiones')

    """
    from apps.modulo_3.cursos.models import ComisionDocente
    
    comision = get_object_or_404(Comision, id_comision=comision_id)
    
    if request.method == 'POST':
        docente_id = request.POST.get('docente_id')
        
        if docente_id:
            try:
                docente_usuario = Usuario.objects.get(id=docente_id)
                # Verificar que el usuario sea realmente un docente
                if Docente.objects.filter(id_persona=docente_usuario.persona).exists():
                    ComisionDocente.objects.get_or_create(
                        fk_id_comision=comision,
                        fk_id_docente=docente_usuario
                    )
                    messages.success(request, f'✅ Docente asignado exitosamente a la comisión #{comision.id_comision}.')
                else:
                    messages.error(request, f'❌ El usuario seleccionado no es un docente válido.')
            except Usuario.DoesNotExist:
                messages.error(request, f'❌ No se encontró el docente seleccionado.')
        else:
            messages.error(request, f'❌ Por favor, selecciona un docente.')
    
    return redirect('administracion:panel_comisiones')
    """


@login_required
@user_passes_test(es_admin)
def panel_inscripciones(request):
    """Panel de gestión de inscripciones con búsqueda y filtros"""
    inscripciones = Inscripcion.objects.all().select_related(
        'estudiante__usuario__persona',
        'comision__fk_id_curso'
    ).order_by('-fecha_hora_inscripcion')
    
    # Búsqueda
    busqueda = request.GET.get('q')
    if busqueda:
        inscripciones = inscripciones.filter(
            Q(estudiante__usuario__persona__nombre__icontains=busqueda) |
            Q(estudiante__usuario__persona__apellido__icontains=busqueda) |
            Q(estudiante__usuario__persona__dni__icontains=busqueda) |
            Q(comision__fk_id_curso__nombre__icontains=busqueda)
        )
    
    # Filtros
    estado_filtro = request.GET.get('estado')
    if estado_filtro:
        inscripciones = inscripciones.filter(estado=estado_filtro)
    
    # Obtener comisiones con cupo disponible para el formulario de inscripción
    # Obtener todas las comisiones abiertas y calcular cupos disponibles usando anotaciones
    comisiones_disponibles = Comision.objects.filter(
        estado='Abierta'
    ).annotate(
        inscritos_count_annotated=Count('inscripciones', filter=Q(inscripciones__estado='confirmado'))
    ).annotate(
        cupos_disponibles_calc=F('cupo_maximo') - F('inscritos_count_annotated')
    ).filter(
        cupos_disponibles_calc__gt=0
    ).select_related('fk_id_curso', 'fk_id_polo').order_by('fk_id_curso__nombre', 'id_comision')
    
    # Obtener estudiantes para el selector (SOLO PRE-INSCRIPTOS como solicitado)
    pre_inscripciones_prefetch = Prefetch(
        'inscripciones',
        queryset=Inscripcion.objects.filter(estado='pre_inscripto').select_related('comision__fk_id_curso'),
        to_attr='pre_inscripciones_list'
    )
    estudiantes = Estudiante.objects.filter(inscripciones__estado='pre_inscripto').distinct().select_related('usuario__persona').prefetch_related(pre_inscripciones_prefetch).order_by('usuario__persona__apellido', 'usuario__persona__nombre')
    
    context = {
        'inscripciones': inscripciones,
        'estado_filtro': estado_filtro,
        'busqueda': busqueda,
        'comisiones_disponibles': comisiones_disponibles,
        'estudiantes': estudiantes,
    }
    return render(request, 'administracion/panel_inscripciones.html', context)


@login_required
@user_passes_test(es_admin)
def inscribir_estudiante_admin(request):
    """Inscribir o confirmar estudiante a una comisión desde el panel de administración"""
    if request.method == 'POST':
        try:
            estudiante_id = request.POST.get('estudiante_id')
            comision_id = request.POST.get('comision_id')
            
            if not estudiante_id or not comision_id:
                messages.error(request, '❌ Por favor, selecciona un estudiante y una comisión.')
                return redirect('administracion:panel_inscripciones')
            
            estudiante = get_object_or_404(Estudiante, pk=estudiante_id)
            comision = get_object_or_404(Comision, id_comision=comision_id)
            
            # Verificar si ya está inscrito
            inscripcion_existente = Inscripcion.objects.filter(estudiante=estudiante, comision=comision).first()
            
            if inscripcion_existente:
                if inscripcion_existente.estado == 'pre_inscripto':
                    # Confirmar inscripción
                    with transaction.atomic():
                        inscripcion_existente.estado = 'confirmado'
                        inscripcion_existente.save()
                    messages.success(request, f'✅ Inscripción confirmada exitosamente para {estudiante.usuario.persona.nombre_completo}.')
                    return redirect('administracion:panel_inscripciones')
                elif inscripcion_existente.estado == 'confirmado':
                    messages.warning(request, f'⚠️ El estudiante {estudiante.usuario.persona.nombre_completo} ya está inscrito y confirmado en esta comisión.')
                    return redirect('administracion:panel_inscripciones')
                elif inscripcion_existente.estado == 'lista_espera':
                    messages.warning(request, f'⚠️ El estudiante {estudiante.usuario.persona.nombre_completo} está en lista de espera.')
                    # Aquí se podría agregar lógica para mover de lista de espera a confirmado si hay cupo
                    return redirect('administracion:panel_inscripciones')
            
            # Si no existe inscripción previa (o se permite crear nueva para otros casos)
            # Verificar cupo disponible
            if comision.cupo_lleno:
                messages.error(request, f'🚫 La comisión {comision.fk_id_curso.nombre} (Comisión #{comision.id_comision}) no tiene cupos disponibles.')
                return redirect('administracion:panel_inscripciones')
            
            # Verificar rango etario
            curso = comision.fk_id_curso
            persona = estudiante.usuario.persona
            edad_real = persona.edad
            
            if edad_real is None:
                if curso.edad_minima or curso.edad_maxima:
                    messages.error(request, f'⚠️ El estudiante {persona.nombre_completo} no tiene fecha de nacimiento registrada y el curso tiene restricciones de edad.')
                    return redirect('administracion:panel_inscripciones')
            else:
                if curso.edad_minima and edad_real < curso.edad_minima:
                    messages.error(request, f'⛔ El estudiante {persona.nombre_completo} ({edad_real} años) no cumple con la edad mínima ({curso.edad_minima} años).')
                    return redirect('administracion:panel_inscripciones')
                
                if curso.edad_maxima and edad_real > curso.edad_maxima:
                    messages.error(request, f'⛔ El estudiante {persona.nombre_completo} ({edad_real} años) supera la edad máxima ({curso.edad_maxima} años).')
                    return redirect('administracion:panel_inscripciones')

            # Crear inscripción
            with transaction.atomic():
                Inscripcion.objects.create(
                    estudiante=estudiante,
                    comision=comision,
                    estado='confirmado'
                )
            
            messages.success(request, f'✅ Estudiante {estudiante.usuario.persona.nombre_completo} inscrito exitosamente en {comision.fk_id_curso.nombre} (Comisión #{comision.id_comision}). Cupos restantes: {comision.cupos_disponibles}')
            
        except Exception as e:
            messages.error(request, f'❌ Error al inscribir estudiante: {str(e)}')
    
    return redirect('administracion:panel_inscripciones')


@login_required
@user_passes_test(es_admin)
def exportar_inscripciones(request):
    """Exportar inscripciones a CSV"""
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="inscripciones_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'
    
    writer = csv.writer(response)
    writer.writerow(['ID', 'Estudiante', 'DNI', 'Curso', 'Comisión', 'Fecha Inscripción', 'Estado', 'Observaciones Salud'])
    
    inscripciones = Inscripcion.objects.all().select_related('estudiante__usuario__persona', 'comision__fk_id_curso')
    
    for insc in inscripciones:
        writer.writerow([
            insc.id,
            insc.estudiante.usuario.persona.nombre_completo,
            insc.estudiante.usuario.persona.dni,
            insc.comision.fk_id_curso.nombre,
            f"#{insc.comision.id_comision}",
            insc.fecha_hora_inscripcion.strftime('%d/%m/%Y %H:%M'),
            insc.get_estado_display(),
            insc.observaciones_salud or ''
        ])
    
    return response


@login_required
@user_passes_test(es_admin)
def buscador_estudiantes(request):
    """Buscador de estudiantes"""
    estudiantes = Estudiante.objects.all().select_related('usuario__persona')
    
    # Búsqueda
    q = request.GET.get('q')
    if q:
        estudiantes = estudiantes.filter(
            Q(usuario__persona__nombre__icontains=q) |
            Q(usuario__persona__apellido__icontains=q) |
            Q(usuario__persona__dni__icontains=q) |
            Q(usuario__persona__correo__icontains=q)
        )
    
    context = {
        'estudiantes': estudiantes,
        'busqueda': q,
    }
    return render(request, 'administracion/buscador_estudiantes.html', context)


@login_required
@user_passes_test(es_admin)
def exportar_estudiantes(request):
    """Exportar estudiantes a CSV"""
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="estudiantes_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'
    
    writer = csv.writer(response)
    writer.writerow(['DNI', 'Nombre', 'Apellido', 'Email', 'Teléfono', 'Edad', 'Ciudad', 'Nivel Estudios', 'Total Inscripciones'])
    
    estudiantes = Estudiante.objects.all().select_related('usuario__persona')
    
    for est in estudiantes:
        persona = est.usuario.persona
        total_insc = est.inscripciones.filter(estado='confirmado').count()
        
        writer.writerow([
            persona.dni,
            persona.nombre,
            persona.apellido,
            persona.correo,
            persona.telefono or '',
            persona.edad or '',
            persona.ciudad_residencia or '',
            est.get_nivel_estudios_display(),
            total_insc
        ])
    
    return response


@login_required
@user_passes_test(es_admin_completo)
def panel_polos(request):
    """Panel de gestión de Polos Creativos"""
    polos = PoloCreativo.objects.all().annotate(
        total_comisiones=Count('comisiones')
    ).order_by('ciudad', 'nombre')
    
    context = {
        'polos': polos,
    }
    return render(request, 'administracion/panel_polos.html', context)


@login_required
@user_passes_test(es_admin_completo)
def crear_polo(request):
    """Crear un nuevo Polo Creativo"""
    if request.method == 'POST':
        try:
            PoloCreativo.objects.create(
                nombre=request.POST.get('nombre'),
                ciudad=request.POST.get('ciudad'),
                direccion=request.POST.get('direccion'),
                telefono=request.POST.get('telefono', ''),
                email=request.POST.get('email', ''),
                activo=True
            )
            messages.success(request, '✅ Polo Creativo creado exitosamente.')
            return redirect('administracion:panel_polos')
        except Exception as e:
            messages.error(request, f'❌ Error al crear polo: {str(e)}')
    
    return render(request, 'administracion/crear_polo.html')


@login_required
@user_passes_test(es_admin)
def estadisticas_detalladas(request):
    """Panel de estadísticas detalladas con gráficos"""
    from apps.modulo_4.asistencia.models import RegistroAsistencia
    from django.db.models import Avg
    
    # Estadísticas generales
    total_cursos = Curso.objects.count()
    total_estudiantes = Estudiante.objects.count()
    total_inscripciones = Inscripcion.objects.filter(estado='confirmado').count()
    
    # Cursos con cantidad de alumnos
    cursos_con_alumnos = Curso.objects.annotate(
        total_alumnos=Count('comision__inscripciones', filter=Q(comision__inscripciones__estado='confirmado'))
    ).order_by('-total_alumnos')
    
    # Datos para gráficos
    datos_grafico = []
    for curso in cursos_con_alumnos:
        if curso.total_alumnos > 0:
            datos_grafico.append({
                'nombre': curso.nombre,
                'cantidad': curso.total_alumnos
            })
    
    # Estadísticas de asistencia
    if RegistroAsistencia.objects.exists():
        promedio_asistencia = RegistroAsistencia.objects.aggregate(Avg('porcentaje_asistencia'))['porcentaje_asistencia__avg'] or 0
    else:
        promedio_asistencia = 0
    
    # Cursos completados vs en proceso
    total_completados = RegistroAsistencia.objects.filter(cumple_requisito_certificado=True).count()
    total_en_proceso = total_inscripciones - total_completados
    
    import json
    
    context = {
        'total_cursos': total_cursos,
        'total_estudiantes': total_estudiantes,
        'total_inscripciones': total_inscripciones,
        'cursos_con_alumnos': cursos_con_alumnos,
        'datos_grafico_json': json.dumps(datos_grafico),
        'promedio_asistencia': round(promedio_asistencia, 2),
        'total_completados': total_completados,
        'total_en_proceso': total_en_proceso,
    }
    return render(request, 'administracion/estadisticas.html', context)


@login_required
@user_passes_test(es_admin)
def api_buscar_estudiantes(request):
    """API para búsqueda en tiempo real de estudiantes"""
    from django.http import JsonResponse
    
    query = request.GET.get('q', '').strip()
    
    if not query or len(query) < 2:
        return JsonResponse({'estudiantes': []})
    
    # Buscar estudiantes
    estudiantes = Estudiante.objects.filter(
        Q(usuario__persona__nombre__icontains=query) |
        Q(usuario__persona__apellido__icontains=query) |
        Q(usuario__persona__dni__icontains=query) |
        Q(usuario__persona__correo__icontains=query)
    ).select_related('usuario__persona')[:20]  # Limitar a 20 resultados
    
    resultados = []
    for est in estudiantes:
        persona = est.usuario.persona
        total_insc = Inscripcion.objects.filter(estudiante=est, estado='confirmado').count()
        
        resultados.append({
            'dni': persona.dni,
            'nombre_completo': persona.nombre_completo,
            'correo': persona.correo,
            'telefono': persona.telefono or 'N/A',
            'edad': persona.edad or 'N/A',
            'ciudad': persona.ciudad_residencia or 'N/A',
            'nivel_estudios': est.get_nivel_estudios_display(),
            'inscripciones': total_insc
        })
    
    return JsonResponse({'estudiantes': resultados})


@login_required
@user_passes_test(es_admin_completo)
def gestion_usuarios(request):
    """Panel de gestión completa de usuarios con buscador"""
    from django.http import JsonResponse
    
    # Función auxiliar para obtener roles de una persona
    def obtener_roles(persona, usuario=None):
        roles_list = []
        if usuario:
            # Verificar Estudiante
            if Estudiante.objects.filter(usuario=usuario).exists():
                roles_list.append('Estudiante')
            # Verificar Administrador
            if UsuarioRol.objects.filter(usuario_id=usuario, rol_id__nombre='Administrador').exists():
                roles_list.append('Administrador')
            # Verificar Mesa de Entrada
            if UsuarioRol.objects.filter(usuario_id=usuario, rol_id__nombre='Mesa de Entrada').exists():
                roles_list.append('Mesa de Entrada')
        # Verificar Docente (usa id_persona directamente)
        if Docente.objects.filter(id_persona=persona).exists():
            roles_list.append('Docente')
        return ', '.join(roles_list) if roles_list else 'Sin rol'
    
    # Si es una petición AJAX, devolver JSON
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        query = request.GET.get('q', '').strip()
        
        # Buscar en Personas (si hay query) o mostrar todos
        if query and len(query) >= 2:
            personas = Persona.objects.filter(
                Q(nombre__icontains=query) |
                Q(apellido__icontains=query) |
                Q(dni__icontains=query) |
                Q(correo__icontains=query)
            ).prefetch_related('usuario_set')[:50]
        else:
            # Si no hay búsqueda, devolver lista vacía para AJAX
            return JsonResponse({'usuarios': []})
        
        resultados = []
        for persona in personas:
            # Obtener el primer usuario asociado (puede haber varios)
            usuario = persona.usuario_set.first()
            
            if usuario:
                roles = obtener_roles(persona, usuario)
                resultados.append({
                    'id': persona.id,
                    'dni': persona.dni,
                    'nombre_completo': persona.nombre_completo,
                    'correo': persona.correo,
                    'telefono': persona.telefono or 'N/A',
                    'edad': persona.edad or 'N/A',
                    'ciudad': persona.ciudad_residencia or 'N/A',
                    'roles': roles,
                    'fecha_nacimiento': persona.fecha_nacimiento.strftime('%Y-%m-%d') if persona.fecha_nacimiento else None,
                })
        
        return JsonResponse({'usuarios': resultados})
    
    # Vista normal - Mostrar todos los usuarios al cargar
    # Obtener todas las personas que tienen al menos un usuario
    personas = Persona.objects.filter(usuario__isnull=False).distinct().prefetch_related('usuario_set').order_by('apellido', 'nombre')[:100]
    
    usuarios_list = []
    for persona in personas:
        # Obtener el primer usuario asociado
        usuario = persona.usuario_set.first()
        
        if usuario:
            roles = obtener_roles(persona, usuario)
            usuarios_list.append({
                'persona': persona,
                'usuario': usuario,
                'roles': roles,
            })
    
    context = {
        'usuarios': usuarios_list,
        'total_usuarios': len(usuarios_list),
        'puede_crear_usuarios': es_admin_completo(request.user)
    }
    return render(request, 'administracion/gestion_usuarios.html', context)


@login_required
@user_passes_test(es_admin_completo)
def crear_usuario_admin(request):
    """Crear un nuevo usuario desde el panel de administración"""
    if request.method == 'POST':
        try:
            with transaction.atomic():
                # Datos personales
                dni = request.POST.get('dni').strip()
                nombre = request.POST.get('nombre').strip()
                apellido = request.POST.get('apellido').strip()
                correo = request.POST.get('correo').strip()
                telefono = request.POST.get('telefono', '').strip()
                fecha_nacimiento = request.POST.get('fecha_nacimiento')
                genero = request.POST.get('genero')
                ciudad = request.POST.get('ciudad_residencia')
                contrasena = request.POST.get('contrasena')
                tipo_usuario = request.POST.get('tipo_usuario')
                
                # Validar que el DNI no exista
                if Persona.objects.filter(dni=dni).exists():
                    messages.error(request, f'❌ El DNI {dni} ya está registrado en el sistema.')
                    return redirect('administracion:gestion_usuarios')
                
                # Crear Persona
                persona = Persona.objects.create(
                    dni=dni,
                    nombre=nombre,
                    apellido=apellido,
                    correo=correo,
                    telefono=telefono,
                    fecha_nacimiento=fecha_nacimiento if fecha_nacimiento else None,
                    genero=genero if genero else None,
                    ciudad_residencia=ciudad if ciudad else None,
                )
                
                # Crear Usuario
                usuario = Usuario.objects.create(
                    persona=persona,
                    contrasena=contrasena,
                )
                
                # Asignar rol según tipo de usuario
                if tipo_usuario == 'estudiante':
                    Estudiante.objects.create(
                        usuario=usuario,
                        nivel_estudios='OT',  # 'OT' = Otro (valor válido del modelo)
                        institucion_actual='Por definir'
                    )
                    rol, _ = Rol.objects.get_or_create(nombre='Estudiante', defaults={'descripcion': 'Rol para estudiantes', 'jerarquia': 3})
                    UsuarioRol.objects.get_or_create(usuario_id=usuario, rol_id=rol)
                    
                elif tipo_usuario == 'docente':
                    Docente.objects.create(
                        id_persona=persona,
                        especialidad='General',
                        experiencia='Docente del sistema'
                    )
                    rol, _ = Rol.objects.get_or_create(nombre='Docente', defaults={'descripcion': 'Rol para docentes', 'jerarquia': 2})
                    UsuarioRol.objects.get_or_create(usuario_id=usuario, rol_id=rol)
                    
                elif tipo_usuario == 'admin':
                    rol, _ = Rol.objects.get_or_create(nombre='Administrador', defaults={'descripcion': 'Rol para administradores', 'jerarquia': 1})
                    UsuarioRol.objects.get_or_create(usuario_id=usuario, rol_id=rol)
                    
                elif tipo_usuario == 'mesa_entrada':
                    rol, _ = Rol.objects.get_or_create(nombre='Mesa de Entrada', defaults={'descripcion': 'Rol para personal de mesa de entrada. Puede gestionar cursos, comisiones, inscripciones y usuarios (excepto crear nuevos usuarios).', 'jerarquia': 2})
                    UsuarioRol.objects.get_or_create(usuario_id=usuario, rol_id=rol)
                
                messages.success(request, f'✅ Usuario {nombre} {apellido} creado exitosamente.')
                return redirect('administracion:gestion_usuarios')
                
        except Exception as e:
            messages.error(request, f'❌ Error al crear usuario: {str(e)}')
            return redirect('administracion:gestion_usuarios')
    
    return render(request, 'administracion/crear_usuario.html')


@login_required
@user_passes_test(es_admin)
def editar_usuario_admin(request, persona_id):
    """Editar un usuario existente"""
    persona = get_object_or_404(Persona, id=persona_id)
    
    # Obtener el primer usuario asociado
    usuario = persona.usuario_set.first()
    if not usuario:
        messages.error(request, '❌ Este usuario no tiene un perfil asociado.')
        return redirect('administracion:gestion_usuarios')
    
    if request.method == 'POST':
        try:
            with transaction.atomic():
                # Actualizar datos personales
                persona.nombre = request.POST.get('nombre', persona.nombre)
                persona.apellido = request.POST.get('apellido', persona.apellido)
                persona.correo = request.POST.get('correo', persona.correo)
                persona.telefono = request.POST.get('telefono', persona.telefono)
                
                fecha_nac = request.POST.get('fecha_nacimiento')
                if fecha_nac:
                    persona.fecha_nacimiento = fecha_nac
                
                persona.genero = request.POST.get('genero', persona.genero)
                persona.ciudad_residencia = request.POST.get('ciudad_residencia', persona.ciudad_residencia)
                persona.save()
                
                # Actualizar contraseña si se proporciona
                nueva_contrasena = request.POST.get('contrasena')
                if nueva_contrasena:
                    usuario.contrasena = nueva_contrasena
                    usuario.save()
                
                # Gestionar roles
                nuevo_rol = request.POST.get('nuevo_rol')
                if nuevo_rol:
                    # Eliminar todos los roles actuales del usuario
                    UsuarioRol.objects.filter(usuario_id=usuario).delete()
                    
                    # Eliminar perfiles existentes
                    Estudiante.objects.filter(usuario=usuario).delete()
                    Docente.objects.filter(id_persona=persona).delete()
                    
                    # Asignar nuevo rol y crear perfil correspondiente
                    if nuevo_rol == 'estudiante':
                        Estudiante.objects.create(
                            usuario=usuario,
                            nivel_estudios='OT',
                            institucion_actual='Por definir'
                        )
                        rol, _ = Rol.objects.get_or_create(nombre='Estudiante', defaults={'descripcion': 'Rol para estudiantes', 'jerarquia': 3})
                        UsuarioRol.objects.get_or_create(usuario_id=usuario, rol_id=rol)
                    elif nuevo_rol == 'docente':
                        Docente.objects.create(
                            id_persona=persona,
                            especialidad='General',
                            experiencia='Docente del sistema'
                        )
                        rol, _ = Rol.objects.get_or_create(nombre='Docente', defaults={'descripcion': 'Rol para docentes', 'jerarquia': 2})
                        UsuarioRol.objects.get_or_create(usuario_id=usuario, rol_id=rol)
                    elif nuevo_rol == 'admin':
                        rol, _ = Rol.objects.get_or_create(nombre='Administrador', defaults={'descripcion': 'Rol para administradores', 'jerarquia': 1})
                        UsuarioRol.objects.get_or_create(usuario_id=usuario, rol_id=rol)
                    elif nuevo_rol == 'mesa_entrada':
                        rol, _ = Rol.objects.get_or_create(nombre='Mesa de Entrada', defaults={'descripcion': 'Rol para personal de mesa de entrada', 'jerarquia': 2})
                        UsuarioRol.objects.get_or_create(usuario_id=usuario, rol_id=rol)
                
                messages.success(request, f'✅ Usuario {persona.nombre_completo} actualizado exitosamente.')
                return redirect('administracion:gestion_usuarios')
                
        except Exception as e:
            messages.error(request, f'❌ Error al actualizar usuario: {str(e)}')
    
    # Determinar roles actuales
    roles_actuales = []
    if Estudiante.objects.filter(usuario=usuario).exists():
        roles_actuales.append('Estudiante')
    if Docente.objects.filter(id_persona=persona).exists():
        roles_actuales.append('Docente')
    if UsuarioRol.objects.filter(usuario_id=usuario, rol_id__nombre='Administrador').exists():
        roles_actuales.append('Administrador')
    if UsuarioRol.objects.filter(usuario_id=usuario, rol_id__nombre='Mesa de Entrada').exists():
        roles_actuales.append('Mesa de Entrada')
    
    context = {
        'persona': persona,
        'usuario': usuario,
        'roles_actuales': ', '.join(roles_actuales) if roles_actuales else 'Sin rol'
    }
    return render(request, 'administracion/editar_usuario.html', context)


@login_required
@user_passes_test(es_admin)
def eliminar_usuario_admin(request, persona_id):
    """Eliminar un usuario del sistema"""
    if request.method == 'POST':
        try:
            persona = get_object_or_404(Persona, id=persona_id)
            nombre_completo = persona.nombre_completo
            
            # Obtener el primer usuario asociado
            usuario = persona.usuario_set.first()
            
            # No permitir eliminar al usuario actual
            if usuario and usuario.persona.dni == request.user.username:
                messages.error(request, '❌ No puedes eliminar tu propio usuario.')
                return redirect('administracion:gestion_usuarios')
            
            with transaction.atomic():
                # Eliminar usuario (cascada eliminará Persona y relaciones)
                if usuario:
                    usuario.delete()
                else:
                    # Si no hay usuario, eliminar solo la persona
                    persona.delete()
            
            messages.success(request, f'✅ Usuario {nombre_completo} eliminado exitosamente.')
            
        except Exception as e:
            messages.error(request, f'❌ Error al eliminar usuario: {str(e)}')
    
    return redirect('administracion:gestion_usuarios')


@login_required
@user_passes_test(es_admin)
def exportar_usuarios_excel(request):
    """Exportar usuarios a Excel"""
    # Crear workbook y worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Usuarios del Sistema"
    
    # Estilos para el encabezado
    header_fill = PatternFill(start_color="6366f1", end_color="6366f1", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Encabezados
    headers = ['DNI', 'Nombre', 'Apellido', 'Email', 'Teléfono', 'Fecha Nacimiento', 
               'Edad', 'Género', 'Ciudad', 'Roles', 'Estado']
    
    # Escribir encabezados con estilo
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Ajustar ancho de columnas
    column_widths = [15, 20, 20, 30, 15, 15, 8, 10, 15, 25, 12]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = width
    
    # Función auxiliar para obtener roles
    def obtener_roles(persona, usuario=None):
        roles_list = []
        if usuario:
            if Estudiante.objects.filter(usuario=usuario).exists():
                roles_list.append('Estudiante')
            if UsuarioRol.objects.filter(usuario_id=usuario, rol_id__nombre='Administrador').exists():
                roles_list.append('Administrador')
        if Docente.objects.filter(id_persona=persona).exists():
            roles_list.append('Docente')
        return ', '.join(roles_list) if roles_list else 'Sin rol'
    
    # Obtener todos los usuarios
    personas = Persona.objects.filter(usuario__isnull=False).distinct().prefetch_related('usuario_set').order_by('apellido', 'nombre')
    
    # Escribir datos
    row_num = 2
    for persona in personas:
        usuario = persona.usuario_set.first()
        if usuario:
            roles = obtener_roles(persona, usuario)
            
            # Datos de la fila
            row_data = [
                persona.dni,
                persona.nombre,
                persona.apellido,
                persona.correo,
                persona.telefono or '',
                persona.fecha_nacimiento.strftime('%d/%m/%Y') if persona.fecha_nacimiento else '',
                persona.edad if persona.edad else '',
                persona.get_genero_display() if persona.genero else '',
                persona.ciudad_residencia or '',
                roles,
                'Activo' if usuario.activo else 'Inactivo'
            ]
            
            # Escribir fila
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                # Alternar color de fondo para mejor legibilidad
                if row_num % 2 == 0:
                    cell.fill = PatternFill(start_color="f8fafc", end_color="f8fafc", fill_type="solid")
            
            row_num += 1
    
    # Crear respuesta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"usuarios_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # Guardar workbook en la respuesta
    wb.save(response)
    
    return response


@login_required
@user_passes_test(es_admin)
def panel_asistencia(request):
    """Panel para seleccionar curso/comisión y ver asistencias"""
    # Verificar si es docente (para filtrar solo sus comisiones)
    es_docente = False
    usuario_actual = None
    comisiones_docente = []
    
    try:
        usuario_actual = Usuario.objects.get(persona__dni=request.user.username)
        if Docente.objects.filter(id_persona=usuario_actual.persona).exists():
            es_docente = True
            from apps.modulo_3.cursos.models import ComisionDocente
            comisiones_docente = ComisionDocente.objects.filter(
                fk_id_docente=usuario_actual
            ).values_list('fk_id_comision_id', flat=True)
    except Usuario.DoesNotExist:
        pass
    
    # Obtener todas las comisiones
    # Si es docente, solo mostrar sus comisiones asignadas
    if es_docente and comisiones_docente:
        comisiones = Comision.objects.filter(
            id_comision__in=comisiones_docente
        ).distinct().select_related('fk_id_curso', 'fk_id_polo').order_by('fk_id_curso__nombre', 'id_comision')
    else:
        comisiones = Comision.objects.all().distinct().select_related('fk_id_curso', 'fk_id_polo').order_by('fk_id_curso__nombre', 'id_comision')
    
    # Si se selecciona una comisión específica, mostrar sus asistencias detalladas y formulario de toma
    comision_id = request.GET.get('comision_id') or request.POST.get('comision_id')
    comision = None
    inscripciones = None
    asistencias_dict = {}
    fechas_clases = []
    fecha_seleccionada = request.GET.get('fecha', date.today().isoformat())
    asistencias_existentes = {}
    
    if comision_id:
        comision = get_object_or_404(Comision, id_comision=comision_id)
        
        # Verificar permisos si es docente
        if es_docente:
             # Verificar que el docente tenga acceso a esta comisión
            from apps.modulo_3.cursos.models import ComisionDocente
            tiene_acceso = ComisionDocente.objects.filter(
                fk_id_docente=usuario_actual,
                fk_id_comision=comision
            ).exists()
            if not tiene_acceso:
                messages.error(request, '❌ No tienes permiso para ver esta comisión.')
                return redirect('administracion:panel_asistencia')

        inscripciones = Inscripcion.objects.filter(
            comision=comision,
            estado='confirmado'
        ).select_related('estudiante__usuario__persona').order_by('estudiante__usuario__persona__apellido')
        
        # PROCESAR POST (Guardar asistencia)
        if request.method == 'POST' and 'guardar_asistencia' in request.POST:
            fecha_clase = request.POST.get('fecha_clase')
            if not fecha_clase:
                 messages.error(request, '❌ Debe seleccionar una fecha.')
            else:
                # Obtener nombre del usuario que registra
                usuario_registro = request.user
                nombre_registrador = f"{usuario_registro.first_name} {usuario_registro.last_name}".strip()
                if not nombre_registrador:
                    nombre_registrador = usuario_registro.username
        
                count_created = 0
                count_updated = 0
        
                try:
                    with transaction.atomic():
                        for inscripcion in inscripciones:
                            presente = request.POST.get(f'presente_{inscripcion.id}') == 'on'
                            observacion = request.POST.get(f'observacion_{inscripcion.id}', '').strip()
                            
                            asistencia, created = Asistencia.objects.update_or_create(
                                inscripcion=inscripcion,
                                fecha_clase=fecha_clase,
                                defaults={
                                    'presente': presente,
                                    'observaciones': observacion,
                                    'registrado_por': nombre_registrador
                                }
                            )
                            
                            if created:
                                count_created += 1
                            else:
                                count_updated += 1
                    
                    messages.success(request, f'✅ Asistencia guardada para el {fecha_clase}. Registros procesados: {count_created + count_updated}.')
                    # Redirigir para limpiar POST
                    from django.urls import reverse
                    return redirect(reverse('administracion:panel_asistencia') + f'?comision_id={comision.id_comision}&fecha={fecha_clase}')
                    
                except Exception as e:
                    messages.error(request, f'❌ Error al guardar asistencia: {str(e)}')
        
        # Obtener asistencias existentes para la fecha seleccionada (para el formulario)
        asistencias_query_fecha = Asistencia.objects.filter(
            inscripcion__in=inscripciones, 
            fecha_clase=fecha_seleccionada
        )
        for a in asistencias_query_fecha:
            asistencias_existentes[a.inscripcion_id] = a

        # Obtener todas las fechas de clases únicas para esta comisión (para el historial)
        fechas_clases = list(Asistencia.objects.filter(
            inscripcion__comision=comision
        ).values_list('fecha_clase', flat=True).distinct().order_by('-fecha_clase'))
        
        # Organizar asistencias por estudiante y fecha (para el historial)
        for inscripcion in inscripciones:
            for fecha in fechas_clases:
                asistencia = Asistencia.objects.filter(
                    inscripcion=inscripcion,
                    fecha_clase=fecha
                ).first()
                key = f"{inscripcion.id}_{fecha}"
                asistencias_dict[key] = asistencia
    
    context = {
        'comisiones': comisiones,
        'comision': comision,
        'inscripciones': inscripciones,
        'asistencias_dict': asistencias_dict,
        'fechas_clases': fechas_clases,
        'es_docente': es_docente,
        'fecha_seleccionada': fecha_seleccionada,
        'asistencias_existentes': asistencias_existentes,
    }
    return render(request, 'administracion/panel_asistencia.html', context)


@login_required
@user_passes_test(es_admin)
def tomar_asistencia_masiva(request, comision_id):
    """Tomar asistencia masiva para una comisión"""
    from django.urls import reverse
    
    comision = get_object_or_404(Comision, id_comision=comision_id)
    
    # Verificar permisos (si es docente)
    es_docente = False
    try:
        usuario_actual = Usuario.objects.get(persona__dni=request.user.username)
        if Docente.objects.filter(id_persona=usuario_actual.persona).exists():
            es_docente = True
            from apps.modulo_3.cursos.models import ComisionDocente
            tiene_acceso = ComisionDocente.objects.filter(
                fk_id_docente=usuario_actual,
                fk_id_comision=comision
            ).exists()
            if not tiene_acceso:
                messages.error(request, '❌ No tienes permiso para gestionar asistencias de esta comisión.')
                return redirect('administracion:panel_asistencia')
    except Usuario.DoesNotExist:
        pass

    inscripciones = Inscripcion.objects.filter(
        comision=comision,
        estado='confirmado'
    ).select_related('estudiante__usuario__persona').order_by('estudiante__usuario__persona__apellido')

    if request.method == 'POST':
        fecha_clase = request.POST.get('fecha_clase')
        if not fecha_clase:
             messages.error(request, '❌ Debe seleccionar una fecha.')
             return redirect(reverse('administracion:tomar_asistencia_masiva', args=[comision_id]))
        
        # Obtener nombre del usuario que registra
        usuario_actual = request.user
        nombre_registrador = f"{usuario_actual.first_name} {usuario_actual.last_name}".strip()
        if not nombre_registrador:
            nombre_registrador = usuario_actual.username

        count_created = 0
        count_updated = 0

        try:
            with transaction.atomic():
                for inscripcion in inscripciones:
                    presente = request.POST.get(f'presente_{inscripcion.id}') == 'on'
                    observacion = request.POST.get(f'observacion_{inscripcion.id}', '').strip()
                    
                    asistencia, created = Asistencia.objects.update_or_create(
                        inscripcion=inscripcion,
                        fecha_clase=fecha_clase,
                        defaults={
                            'presente': presente,
                            'observaciones': observacion,
                            'registrado_por': nombre_registrador
                        }
                    )
                    
                    if created:
                        count_created += 1
                    else:
                        count_updated += 1
            
            messages.success(request, f'✅ Asistencia guardada para el {fecha_clase}. Registros procesados: {count_created + count_updated}.')
            return redirect(reverse('administracion:panel_asistencia') + f'?comision_id={comision.id_comision}')
            
        except Exception as e:
            messages.error(request, f'❌ Error al guardar asistencia: {str(e)}')
            return redirect(reverse('administracion:tomar_asistencia_masiva', args=[comision_id]) + f'?fecha={fecha_clase}')

    # GET
    fecha_param = request.GET.get('fecha', date.today().isoformat())
    
    # Obtener asistencias existentes para la fecha seleccionada
    asistencias_existentes = {}
    asistencias_query = Asistencia.objects.filter(
        inscripcion__in=inscripciones, 
        fecha_clase=fecha_param
    )
    for a in asistencias_query:
        asistencias_existentes[a.inscripcion_id] = a
    
    context = {
        'comision': comision,
        'inscripciones': inscripciones,
        'fecha_seleccionada': fecha_param,
        'asistencias_existentes': asistencias_existentes,
        'es_docente': es_docente,
    }
    return render(request, 'administracion/tomar_asistencia_masiva.html', context)


@login_required
@user_passes_test(es_admin)
def crear_editar_asistencia(request, inscripcion_id):
    """Crear o editar asistencia de un estudiante"""
    inscripcion = get_object_or_404(Inscripcion, id=inscripcion_id, estado='confirmado')
    
    # Verificar si es docente
    es_docente = False
    try:
        usuario_actual = Usuario.objects.get(persona__dni=request.user.username)
        if Docente.objects.filter(id_persona=usuario_actual.persona).exists():
            es_docente = True
            # Verificar que el docente tenga acceso a esta comisión
            from apps.modulo_3.cursos.models import ComisionDocente
            tiene_acceso = ComisionDocente.objects.filter(
                fk_id_docente=usuario_actual,
                fk_id_comision=inscripcion.comision
            ).exists()
            if not tiene_acceso:
                messages.error(request, '❌ No tienes permiso para gestionar asistencias de esta comisión.')
                return redirect('administracion:panel_asistencia')
    except Usuario.DoesNotExist:
        pass
    
    if request.method == 'POST':
        try:
            fecha_clase = request.POST.get('fecha_clase')
            presente = request.POST.get('presente') == 'on'
            observaciones = request.POST.get('observaciones', '').strip()
            
            if not fecha_clase:
                messages.error(request, '❌ Por favor, selecciona una fecha.')
                from django.urls import reverse
                return redirect(reverse('administracion:panel_asistencia') + f'?comision_id={inscripcion.comision.id_comision}')
            
            # Obtener nombre del usuario que registra
            usuario_actual = request.user
            nombre_registrador = f"{usuario_actual.first_name} {usuario_actual.last_name}".strip()
            if not nombre_registrador:
                nombre_registrador = usuario_actual.username
            
            # Crear o actualizar asistencia
            asistencia, created = Asistencia.objects.update_or_create(
                inscripcion=inscripcion,
                fecha_clase=fecha_clase,
                defaults={
                    'presente': presente,
                    'observaciones': observaciones,
                    'registrado_por': nombre_registrador,
                }
            )
            
            # El RegistroAsistencia se actualiza automáticamente mediante señales (signals)
            # No es necesario actualizarlo manualmente aquí
            
            if created:
                messages.success(request, f'✅ Asistencia registrada exitosamente para {inscripcion.estudiante.usuario.persona.nombre_completo}. El avance se ha actualizado automáticamente.')
            else:
                messages.success(request, f'✅ Asistencia actualizada exitosamente para {inscripcion.estudiante.usuario.persona.nombre_completo}. El avance se ha actualizado automáticamente.')
            
            from django.urls import reverse
            return redirect(reverse('administracion:panel_asistencia') + f'?comision_id={inscripcion.comision.id_comision}')
            
        except Exception as e:
            messages.error(request, f'❌ Error al registrar asistencia: {str(e)}')
    
    # GET: mostrar formulario
    fecha_clase_param = request.GET.get('fecha_clase', date.today().isoformat())
    asistencia_existente = Asistencia.objects.filter(
        inscripcion=inscripcion,
        fecha_clase=fecha_clase_param
    ).first()
    
    context = {
        'inscripcion': inscripcion,
        'asistencia': asistencia_existente,
        'fecha_clase': fecha_clase_param,
        'es_docente': es_docente,
    }
    return render(request, 'administracion/crear_editar_asistencia.html', context)


@login_required
@user_passes_test(es_admin)
def eliminar_asistencia(request, asistencia_id):
    """Eliminar una asistencia"""
    asistencia = get_object_or_404(Asistencia, id_asistencia=asistencia_id)
    comision_id = asistencia.inscripcion.comision.id_comision
    estudiante_nombre = asistencia.inscripcion.estudiante.usuario.persona.nombre_completo
    inscripcion = asistencia.inscripcion  # Guardar referencia antes de eliminar
    
    # Verificar si es docente
    es_docente = False
    try:
        usuario_actual = Usuario.objects.get(persona__dni=request.user.username)
        if Docente.objects.filter(id_persona=usuario_actual.persona).exists():
            es_docente = True
            # Verificar que el docente tenga acceso a esta comisión
            from apps.modulo_3.cursos.models import ComisionDocente
            tiene_acceso = ComisionDocente.objects.filter(
                fk_id_docente=usuario_actual,
                fk_id_comision=inscripcion.comision
            ).exists()
            if not tiene_acceso:
                messages.error(request, '❌ No tienes permiso para eliminar asistencias de esta comisión.')
                return redirect('administracion:panel_asistencia')
    except Usuario.DoesNotExist:
        pass
    
    if request.method == 'POST':
        try:
            asistencia.delete()
            # El registro se actualizará automáticamente mediante la señal post_delete
            messages.success(request, f'✅ Asistencia eliminada exitosamente para {estudiante_nombre}. El avance se ha actualizado automáticamente.')
        except Exception as e:
            messages.error(request, f'❌ Error al eliminar asistencia: {str(e)}')
        
        from django.urls import reverse
        return redirect(reverse('administracion:panel_asistencia') + f'?comision_id={comision_id}')
    
    context = {
        'asistencia': asistencia,
        'es_docente': es_docente,
    }
    return render(request, 'administracion/eliminar_asistencia.html', context)


@login_required
@user_passes_test(es_admin)
def exportar_estadisticas_estudiantes_curso(request):
    """Exportar estadísticas de estudiantes por curso a Excel"""
    from apps.modulo_4.asistencia.models import RegistroAsistencia
    from django.db.models import Avg
    
    # Crear workbook
    wb = Workbook()
    
    # Hoja 1: Resumen por Curso
    ws1 = wb.active
    ws1.title = "Estudiantes por Curso"
    
    # Estilos
    header_fill = PatternFill(start_color="6366f1", end_color="6366f1", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Encabezados
    headers = ['Curso', 'Total Alumnos', 'Completados', 'En Proceso', '% Completados']
    for col_num, header in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Ajustar ancho de columnas
    column_widths = [40, 15, 15, 15, 15]
    for col_num, width in enumerate(column_widths, 1):
        ws1.column_dimensions[get_column_letter(col_num)].width = width
    
    # Obtener cursos con estadísticas
    cursos_con_alumnos = Curso.objects.annotate(
        total_alumnos=Count('comision__inscripciones', filter=Q(comision__inscripciones__estado='confirmado'))
    ).filter(total_alumnos__gt=0).order_by('-total_alumnos')
    
    # Escribir datos
    row_num = 2
    for curso in cursos_con_alumnos:
        # Obtener inscripciones del curso
        inscripciones_curso = Inscripcion.objects.filter(
            comision__fk_id_curso=curso,
            estado='confirmado'
        )
        
        # Contar completados (con RegistroAsistencia que cumple requisito)
        completados = RegistroAsistencia.objects.filter(
            inscripcion__in=inscripciones_curso,
            cumple_requisito_certificado=True
        ).count()
        
        en_proceso = inscripciones_curso.count() - completados
        porcentaje = (completados / curso.total_alumnos * 100) if curso.total_alumnos > 0 else 0
        
        row_data = [
            curso.nombre,
            curso.total_alumnos,
            completados,
            en_proceso,
            f"{porcentaje:.2f}%"
        ]
        
        for col_num, value in enumerate(row_data, 1):
            cell = ws1.cell(row=row_num, column=col_num, value=value)
            if row_num % 2 == 0:
                cell.fill = PatternFill(start_color="f8fafc", end_color="f8fafc", fill_type="solid")
        
        row_num += 1
    
    # Hoja 2: Detalle de Estudiantes por Curso
    ws2 = wb.create_sheet("Detalle Estudiantes")
    
    headers_detalle = ['Curso', 'Comisión', 'Estudiante', 'DNI', 'Estado', '% Asistencia']
    for col_num, header in enumerate(headers_detalle, 1):
        cell = ws2.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    column_widths_detalle = [30, 10, 30, 15, 15, 15]
    for col_num, width in enumerate(column_widths_detalle, 1):
        ws2.column_dimensions[get_column_letter(col_num)].width = width
    
    row_num = 2
    for curso in cursos_con_alumnos:
        inscripciones_curso = Inscripcion.objects.filter(
            comision__fk_id_curso=curso,
            estado='confirmado'
        ).select_related('estudiante__usuario__persona', 'comision')
        
        for inscripcion in inscripciones_curso:
            registro = RegistroAsistencia.objects.filter(inscripcion=inscripcion).first()
            estado = "Completado" if registro and registro.cumple_requisito_certificado else "En Proceso"
            porcentaje_asist = registro.porcentaje_asistencia if registro else 0
            
            row_data = [
                curso.nombre,
                inscripcion.comision.id_comision,
                inscripcion.estudiante.usuario.persona.nombre_completo,
                inscripcion.estudiante.usuario.persona.dni,
                estado,
                f"{porcentaje_asist:.2f}%"
            ]
            
            for col_num, value in enumerate(row_data, 1):
                cell = ws2.cell(row=row_num, column=col_num, value=value)
                if row_num % 2 == 0:
                    cell.fill = PatternFill(start_color="f8fafc", end_color="f8fafc", fill_type="solid")
            
            row_num += 1
    
    # Crear respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"estadisticas_estudiantes_curso_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response


@login_required
@user_passes_test(es_admin)
def exportar_asistencias_por_curso(request):
    """Exportar asistencias agrupadas por curso a Excel"""
    curso_id = request.GET.get('curso_id')
    
    if not curso_id:
        messages.error(request, '❌ Por favor, selecciona un curso.')
        return redirect('administracion:panel_asistencia')
    
    curso = get_object_or_404(Curso, id_curso=curso_id)
    
    # Obtener todas las comisiones del curso
    comisiones = Comision.objects.filter(
        fk_id_curso=curso,
        inscripciones__estado='confirmado'
    ).distinct()
    
    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = f"Asistencias - {curso.nombre[:30]}"
    
    # Estilos
    header_fill = PatternFill(start_color="6366f1", end_color="6366f1", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Encabezados
    headers = ['Comisión', 'Estudiante', 'DNI', 'Fecha Clase', 'Presente', 'Observaciones', 'Registrado por']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Ajustar ancho de columnas
    column_widths = [12, 30, 15, 15, 10, 30, 20]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = width
    
    # Escribir datos
    row_num = 2
    for comision in comisiones:
        inscripciones = Inscripcion.objects.filter(
            comision=comision,
            estado='confirmado'
        ).select_related('estudiante__usuario__persona').prefetch_related(
            Prefetch('asistencias', queryset=Asistencia.objects.order_by('-fecha_clase'))
        )
        
        for inscripcion in inscripciones:
            asistencias = inscripcion.asistencias.all()
            
            for asistencia in asistencias:
                row_data = [
                    comision.id_comision,
                    inscripcion.estudiante.usuario.persona.nombre_completo,
                    inscripcion.estudiante.usuario.persona.dni,
                    asistencia.fecha_clase.strftime('%d/%m/%Y'),
                    'Sí' if asistencia.presente else 'No',
                    asistencia.observaciones or '',
                    asistencia.registrado_por or ''
                ]
                
                for col_num, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_num, column=col_num, value=value)
                    if row_num % 2 == 0:
                        cell.fill = PatternFill(start_color="f8fafc", end_color="f8fafc", fill_type="solid")
                
                row_num += 1
    
    # Crear respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"asistencias_curso_{curso.nombre[:20]}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response


@login_required
@user_passes_test(es_admin)
def exportar_asistencias_por_comision(request):
    """Exportar asistencias de una comisión específica a Excel"""
    comision_id = request.GET.get('comision_id')
    
    if not comision_id:
        messages.error(request, '❌ Por favor, selecciona una comisión.')
        return redirect('administracion:panel_asistencia')
    
    comision = get_object_or_404(Comision, id_comision=comision_id)
    
    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = f"Comisión {comision.id_comision}"
    
    # Estilos
    header_fill = PatternFill(start_color="6366f1", end_color="6366f1", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Información de la comisión
    ws.cell(row=1, column=1, value="Curso:").font = Font(bold=True)
    ws.cell(row=1, column=2, value=comision.fk_id_curso.nombre)
    ws.cell(row=2, column=1, value="Comisión:").font = Font(bold=True)
    ws.cell(row=2, column=2, value=f"#{comision.id_comision}")
    if comision.fk_id_polo:
        ws.cell(row=3, column=1, value="Polo:").font = Font(bold=True)
        ws.cell(row=3, column=2, value=comision.fk_id_polo.nombre)
    
    # Encabezados (fila 5)
    headers = ['Estudiante', 'DNI', 'Fecha Clase', 'Presente', 'Observaciones', 'Registrado por', 'Fecha Registro']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Ajustar ancho de columnas
    column_widths = [30, 15, 15, 10, 30, 20, 20]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = width
    
    # Obtener inscripciones y asistencias
    inscripciones = Inscripcion.objects.filter(
        comision=comision,
        estado='confirmado'
    ).select_related('estudiante__usuario__persona').prefetch_related(
        Prefetch('asistencias', queryset=Asistencia.objects.order_by('-fecha_clase'))
    ).order_by('estudiante__usuario__persona__apellido')
    
    # Escribir datos
    row_num = 6
    for inscripcion in inscripciones:
        asistencias = inscripcion.asistencias.all()
        
        for asistencia in asistencias:
            row_data = [
                inscripcion.estudiante.usuario.persona.nombre_completo,
                inscripcion.estudiante.usuario.persona.dni,
                asistencia.fecha_clase.strftime('%d/%m/%Y'),
                'Sí' if asistencia.presente else 'No',
                asistencia.observaciones or '',
                asistencia.registrado_por or '',
                asistencia.fecha_registro.strftime('%d/%m/%Y %H:%M') if asistencia.fecha_registro else ''
            ]
            
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                if row_num % 2 == 0:
                    cell.fill = PatternFill(start_color="f8fafc", end_color="f8fafc", fill_type="solid")
            
            row_num += 1
    
    # Crear respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"asistencias_comision_{comision.id_comision}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response


# ==================== VISTAS PARA DOCENTES ====================

@login_required
@user_passes_test(es_admin)
def mis_cursos_docente(request):
    """Vista para que el docente vea sus cursos asignados"""
    # DESHABILITADO TEMPORALMENTE
    messages.warning(request, 'La gestión de cursos para docentes no está habilitada por el momento.')
    return redirect('dashboard')

    try:
        usuario = Usuario.objects.get(persona__dni=request.user.username)
        
        # Verificar si es docente
        if not Docente.objects.filter(id_persona=usuario.persona).exists():
            messages.error(request, '❌ No tienes perfil de docente.')
            return redirect('dashboard')
        
        # Obtener comisiones asignadas al docente
        comisiones_asignadas = ComisionDocente.objects.filter(
            fk_id_docente=usuario
        ).select_related('fk_id_comision__fk_id_curso', 'fk_id_comision__fk_id_polo').order_by('fk_id_comision__fk_id_curso__nombre')
        
        # Agrupar por curso
        cursos_dict = {}
        for comision_docente in comisiones_asignadas:
            comision = comision_docente.fk_id_comision
            curso = comision.fk_id_curso
            
            if curso.id_curso not in cursos_dict:
                cursos_dict[curso.id_curso] = {
                    'curso': curso,
                    'comisiones': []
                }
            
            # Contar estudiantes inscritos
            total_estudiantes = Inscripcion.objects.filter(
                comision=comision,
                estado='confirmado'
            ).count()
            
            cursos_dict[curso.id_curso]['comisiones'].append({
                'comision': comision,
                'total_estudiantes': total_estudiantes,
            })
        
        context = {
            'cursos_dict': cursos_dict,
            'es_docente': True,
        }
        return render(request, 'administracion/mis_cursos_docente.html', context)
    except Usuario.DoesNotExist:
        messages.error(request, '❌ No se encontró tu perfil de usuario.')
        return redirect('dashboard')


@login_required
@user_passes_test(es_admin)
def estudiantes_comision(request, comision_id):
    """Vista para que el docente vea los estudiantes de una comisión específica"""
    # DESHABILITADO TEMPORALMENTE
    messages.warning(request, 'La gestión de estudiantes para docentes no está habilitada por el momento.')
    return redirect('dashboard')

    try:
        usuario = Usuario.objects.get(persona__dni=request.user.username)
        comision = get_object_or_404(Comision, id_comision=comision_id)
        
        # Verificar que el docente tenga acceso a esta comisión
        tiene_acceso = ComisionDocente.objects.filter(
            fk_id_docente=usuario,
            fk_id_comision=comision
        ).exists()
        
        if not tiene_acceso:
            messages.error(request, '❌ No tienes permiso para ver esta comisión.')
            return redirect('administracion:mis_cursos_docente')
        
        # Obtener estudiantes inscritos confirmados
        inscripciones = Inscripcion.objects.filter(
            comision=comision,
            estado='confirmado'
        ).select_related('estudiante__usuario__persona').order_by('estudiante__usuario__persona__apellido', 'estudiante__usuario__persona__nombre')
        
        # Calcular asistencias para cada estudiante
        estudiantes_con_asistencia = []
        for inscripcion in inscripciones:
            total_asistencias = Asistencia.objects.filter(inscripcion=inscripcion).count()
            asistencias_presentes = Asistencia.objects.filter(inscripcion=inscripcion, presente=True).count()
            porcentaje = (asistencias_presentes / total_asistencias * 100) if total_asistencias > 0 else 0
            
            estudiantes_con_asistencia.append({
                'inscripcion': inscripcion,
                'total_asistencias': total_asistencias,
                'asistencias_presentes': asistencias_presentes,
                'porcentaje': round(porcentaje, 2),
            })
        
        context = {
            'comision': comision,
            'estudiantes': estudiantes_con_asistencia,
            'es_docente': True,
        }
        return render(request, 'administracion/estudiantes_comision.html', context)
    except Usuario.DoesNotExist:
        messages.error(request, '❌ No se encontró tu perfil de usuario.')
        return redirect('dashboard')


@login_required
@user_passes_test(es_admin)
def materiales_comision(request, comision_id):
    """Vista para que el docente vea y gestione materiales de una comisión"""
    # DESHABILITADO TEMPORALMENTE
    messages.warning(request, 'La gestión de materiales para docentes no está habilitada por el momento.')
    return redirect('dashboard')

    try:
        usuario = Usuario.objects.get(persona__dni=request.user.username)
        comision = get_object_or_404(Comision, id_comision=comision_id)
        
        # Verificar que el docente tenga acceso a esta comisión
        tiene_acceso = ComisionDocente.objects.filter(
            fk_id_docente=usuario,
            fk_id_comision=comision
        ).exists()
        
        if not tiene_acceso:
            messages.error(request, '❌ No tienes permiso para gestionar materiales de esta comisión.')
            return redirect('administracion:mis_cursos_docente')
        
        # Obtener materiales de la comisión
        materiales = Material.objects.filter(
            fk_id_comision=comision
        ).order_by('-fecha_subida')
        
        context = {
            'comision': comision,
            'materiales': materiales,
            'es_docente': True,
        }
        return render(request, 'administracion/materiales_comision.html', context)
    except Usuario.DoesNotExist:
        messages.error(request, '❌ No se encontró tu perfil de usuario.')
        return redirect('dashboard')


@login_required
@user_passes_test(es_admin)
def subir_material(request, comision_id):
    """Vista para que el docente suba un material a una comisión"""
    # DESHABILITADO TEMPORALMENTE
    messages.warning(request, 'La subida de materiales no está habilitada por el momento.')
    return redirect('dashboard')

    try:
        usuario = Usuario.objects.get(persona__dni=request.user.username)
        comision = get_object_or_404(Comision, id_comision=comision_id)
        
        # Verificar que el docente tenga acceso a esta comisión
        tiene_acceso = ComisionDocente.objects.filter(
            fk_id_docente=usuario,
            fk_id_comision=comision
        ).exists()
        
        if not tiene_acceso:
            messages.error(request, '❌ No tienes permiso para subir materiales a esta comisión.')
            return redirect('administracion:mis_cursos_docente')
        
        if request.method == 'POST':
            form = MaterialForm(request.POST, request.FILES)
            if form.is_valid():
                material = form.save(commit=False)
                material.fk_id_comision = comision
                material.fk_id_docente = usuario
                material.save()
                messages.success(request, f'✅ Material "{material.nombre_archivo}" subido exitosamente.')
                return redirect('administracion:materiales_comision', comision_id=comision_id)
        else:
            form = MaterialForm()
        
        context = {
            'comision': comision,
            'form': form,
            'es_docente': True,
        }
        return render(request, 'administracion/subir_material.html', context)
    except Usuario.DoesNotExist:
        messages.error(request, '❌ No se encontró tu perfil de usuario.')
        return redirect('dashboard')


@login_required
@user_passes_test(es_admin)
def eliminar_material(request, material_id):
    """Vista para que el docente elimine un material"""
    # DESHABILITADO TEMPORALMENTE
    messages.warning(request, 'La eliminación de materiales no está habilitada por el momento.')
    return redirect('dashboard')

    try:
        usuario = Usuario.objects.get(persona__dni=request.user.username)
        material = get_object_or_404(Material, id_material=material_id)
        comision_id = material.fk_id_comision.id_comision
        
        # Verificar que el docente tenga acceso a esta comisión y que sea el dueño del material
        tiene_acceso = ComisionDocente.objects.filter(
            fk_id_docente=usuario,
            fk_id_comision=material.fk_id_comision
        ).exists()
        
        if not tiene_acceso or material.fk_id_docente != usuario:
            messages.error(request, '❌ No tienes permiso para eliminar este material.')
            return redirect('administracion:materiales_comision', comision_id=comision_id)
        
        if request.method == 'POST':
            nombre_material = material.nombre_archivo
            material.delete()
            messages.success(request, f'✅ Material "{nombre_material}" eliminado exitosamente.')
            return redirect('administracion:materiales_comision', comision_id=comision_id)
        
        context = {
            'material': material,
            'es_docente': True,
        }
        return render(request, 'administracion/eliminar_material.html', context)
    except Usuario.DoesNotExist:
        messages.error(request, '❌ No se encontró tu perfil de usuario.')
        return redirect('dashboard')


@login_required
@user_passes_test(es_admin)
def estudiantes_docente(request):
    """Vista para que el docente vea todas sus comisiones y pueda seleccionar una para ver estudiantes"""
    # DESHABILITADO TEMPORALMENTE
    messages.warning(request, 'La gestión de estudiantes para docentes no está habilitada por el momento.')
    return redirect('dashboard')

    try:
        usuario = Usuario.objects.get(persona__dni=request.user.username)
        
        # Verificar si es docente
        if not Docente.objects.filter(id_persona=usuario.persona).exists():
            messages.error(request, '❌ No tienes perfil de docente.')
            return redirect('dashboard')
        
        # Obtener comisiones asignadas al docente
        comisiones_asignadas = ComisionDocente.objects.filter(
            fk_id_docente=usuario
        ).select_related('fk_id_comision__fk_id_curso', 'fk_id_comision__fk_id_polo').order_by('fk_id_comision__fk_id_curso__nombre')
        
        # Agrupar por curso y contar estudiantes
        cursos_dict = {}
        for comision_docente in comisiones_asignadas:
            comision = comision_docente.fk_id_comision
            curso = comision.fk_id_curso
            
            if curso.id_curso not in cursos_dict:
                cursos_dict[curso.id_curso] = {
                    'curso': curso,
                    'comisiones': []
                }
            
            # Contar estudiantes inscritos
            total_estudiantes = Inscripcion.objects.filter(
                comision=comision,
                estado='confirmado'
            ).count()
            
            cursos_dict[curso.id_curso]['comisiones'].append({
                'comision': comision,
                'total_estudiantes': total_estudiantes,
            })
        
        context = {
            'cursos_dict': cursos_dict,
            'es_docente': True,
            'vista_estudiantes': True,  # Flag para identificar que es la vista de estudiantes
        }
        return render(request, 'administracion/estudiantes_docente.html', context)
    except Usuario.DoesNotExist:
        messages.error(request, '❌ No se encontró tu perfil de usuario.')
        return redirect('dashboard')


@login_required
@user_passes_test(es_admin)
def materiales_docente(request):
    """Vista para que el docente vea todas sus comisiones y pueda seleccionar una para ver materiales"""
    # DESHABILITADO TEMPORALMENTE
    messages.warning(request, 'La gestión de materiales para docentes no está habilitada por el momento.')
    return redirect('dashboard')

    try:
        usuario = Usuario.objects.get(persona__dni=request.user.username)
        
        # Verificar si es docente
        if not Docente.objects.filter(id_persona=usuario.persona).exists():
            messages.error(request, '❌ No tienes perfil de docente.')
            return redirect('dashboard')
        
        # Obtener comisiones asignadas al docente
        comisiones_asignadas = ComisionDocente.objects.filter(
            fk_id_docente=usuario
        ).select_related('fk_id_comision__fk_id_curso', 'fk_id_comision__fk_id_polo').order_by('fk_id_comision__fk_id_curso__nombre')
        
        # Agrupar por curso y contar materiales
        cursos_dict = {}
        for comision_docente in comisiones_asignadas:
            comision = comision_docente.fk_id_comision
            curso = comision.fk_id_curso
            
            if curso.id_curso not in cursos_dict:
                cursos_dict[curso.id_curso] = {
                    'curso': curso,
                    'comisiones': []
                }
            
            # Contar materiales
            total_materiales = Material.objects.filter(
                fk_id_comision=comision
            ).count()
            
            cursos_dict[curso.id_curso]['comisiones'].append({
                'comision': comision,
                'total_materiales': total_materiales,
            })
        
        context = {
            'cursos_dict': cursos_dict,
            'es_docente': True,
            'vista_materiales': True,  # Flag para identificar que es la vista de materiales
        }
        return render(request, 'administracion/materiales_docente.html', context)
    except Usuario.DoesNotExist:
        messages.error(request, '❌ No se encontró tu perfil de usuario.')
        return redirect('dashboard')


@login_required
@user_passes_test(es_admin)
def docentes_cursos(request):
    """Vista para ver todos los docentes con sus cursos asignados (para Admin y Mesa de Entrada)"""
    # DESHABILITADO TEMPORALMENTE
    messages.warning(request, 'La visualización de docentes y cursos no está habilitada por el momento.')
    return redirect('dashboard_admin')

    try:
        # Obtener todos los docentes
        docentes = Docente.objects.all().select_related('id_persona')
        
        # Agrupar docentes con sus comisiones asignadas
        docentes_con_cursos = []
        for docente in docentes:
            persona = docente.id_persona
            # Buscar el usuario asociado al docente
            try:
                usuario = Usuario.objects.get(persona=persona)
            except Usuario.DoesNotExist:
                continue
            
            # Obtener comisiones asignadas al docente
            comisiones_asignadas = ComisionDocente.objects.filter(
                fk_id_docente=usuario
            ).select_related('fk_id_comision__fk_id_curso', 'fk_id_comision__fk_id_polo').order_by('fk_id_comision__fk_id_curso__nombre')
            
            if comisiones_asignadas.exists():
                # Agrupar por curso
                cursos_dict = {}
                for comision_docente in comisiones_asignadas:
                    comision = comision_docente.fk_id_comision
                    curso = comision.fk_id_curso
                    
                    if curso.id_curso not in cursos_dict:
                        cursos_dict[curso.id_curso] = {
                            'curso': curso,
                            'comisiones': []
                        }
                    
                    # Obtener estudiantes inscritos con sus datos completos
                    inscripciones = Inscripcion.objects.filter(
                        comision=comision,
                        estado='confirmado'
                    ).select_related(
                        'estudiante__usuario__persona'
                    ).order_by('estudiante__usuario__persona__apellido', 'estudiante__usuario__persona__nombre')
                    
                    estudiantes_inscriptos = []
                    for inscripcion in inscripciones:
                        estudiantes_inscriptos.append({
                            'nombre_completo': inscripcion.estudiante.usuario.persona.nombre_completo,
                            'dni': inscripcion.estudiante.usuario.persona.dni,
                            'correo': inscripcion.estudiante.usuario.persona.correo,
                            'telefono': inscripcion.estudiante.usuario.persona.telefono or 'No especificado',
                        })
                    
                    cursos_dict[curso.id_curso]['comisiones'].append({
                        'comision': comision,
                        'total_estudiantes': inscripciones.count(),
                        'estudiantes_inscriptos': estudiantes_inscriptos,
                    })
                
                docentes_con_cursos.append({
                    'docente': docente,
                    'persona': persona,
                    'usuario': usuario,
                    'cursos_dict': cursos_dict,
                    'total_comisiones': comisiones_asignadas.count(),
                })
        
        context = {
            'docentes_con_cursos': docentes_con_cursos,
        }
        return render(request, 'administracion/docentes_cursos.html', context)
    except Exception as e:
        messages.error(request, f'❌ Error al cargar la información: {str(e)}')
        return redirect('dashboard_admin')
