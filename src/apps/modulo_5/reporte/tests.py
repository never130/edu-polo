from datetime import date
from io import BytesIO

from django.test import TestCase
from django.urls import reverse
from openpyxl import load_workbook

from apps.modulo_1.roles.models import Rol, UsuarioRol
from apps.modulo_1.usuario.models import Persona, Usuario
from apps.modulo_5.reporte.apps import ReporteConfig


class ReporteSmokeTests(TestCase):
    def test_app_config_nombre(self):
        self.assertEqual(ReporteConfig.name, 'apps.modulo_5.reporte')


class ReporteExportTests(TestCase):
    def setUp(self):
        self.password = 'pw'
        self.persona = Persona.objects.create(
            dni='85000000',
            nombre='Ana',
            apellido='Admin',
            correo='admin@test.com',
            fecha_nacimiento=date(2000, 1, 1),
            ciudad_residencia='Ushuaia',
        )
        self.usuario = Usuario.objects.create(persona=self.persona, contrasena=self.password)
        rol = Rol.objects.create(nombre='Administrador', descripcion='Admin', jerarquia=1)
        UsuarioRol.objects.create(usuario_id=self.usuario, rol_id=rol)
        self.assertTrue(self.client.login(username=self.persona.dni, password=self.password))

    def test_exportar_usuarios_excel_contiene_hoja_y_encabezados(self):
        response = self.client.get(reverse('administracion:exportar_usuarios_excel'), secure=True)
        self.assertEqual(response.status_code, 200)
        self.assertIn('application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', response['Content-Type'])
        self.assertTrue(response.content.startswith(b'PK'))

        wb = load_workbook(filename=BytesIO(response.content))
        self.assertIn('Usuarios del Sistema', wb.sheetnames)
        ws = wb['Usuarios del Sistema']
        self.assertEqual(ws.cell(row=1, column=1).value, 'DNI')
        self.assertEqual(ws.cell(row=2, column=1).value, self.persona.dni)
