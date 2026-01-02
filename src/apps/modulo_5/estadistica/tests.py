from datetime import date

from io import BytesIO

from django.test import TestCase
from django.urls import reverse
from openpyxl import load_workbook

from apps.modulo_1.roles.models import Rol, UsuarioRol
from apps.modulo_1.usuario.models import Persona, Usuario
from apps.modulo_2.inscripciones.models import Inscripcion
from apps.modulo_3.cursos.models import Comision, Curso, PoloCreativo
from apps.modulo_4.asistencia.models import Asistencia, RegistroAsistencia
from apps.modulo_1.roles.models import Estudiante


class EstadisticaAdminTests(TestCase):
    def setUp(self):
        self.password = 'pw'
        persona = Persona.objects.create(
            dni='84000000',
            nombre='Ana',
            apellido='Admin',
            correo='admin@test.com',
            fecha_nacimiento=date(2000, 1, 1),
            ciudad_residencia='Ushuaia',
        )
        self.usuario = Usuario.objects.create(persona=persona, contrasena=self.password)
        rol = Rol.objects.create(nombre='Administrador', descripcion='Admin', jerarquia=1)
        UsuarioRol.objects.create(usuario_id=self.usuario, rol_id=rol)
        self.assertTrue(self.client.login(username=persona.dni, password=self.password))

    def test_estadisticas_detalladas_responde_200(self):
        response = self.client.get(reverse('administracion:estadisticas'), secure=True)
        self.assertEqual(response.status_code, 200)

    def test_exportar_estadisticas_estudiantes_curso_xlsx(self):
        polo = PoloCreativo.objects.create(nombre='Polo', ciudad='Ushuaia', direccion='Test', activo=True)
        curso = Curso.objects.create(nombre='Curso Stats', estado='Abierto', orden=10)
        comision = Comision.objects.create(
            fk_id_curso=curso,
            fk_id_polo=polo,
            dias_horarios='Miércoles 10:00 - 12:00',
            fecha_inicio=date(2025, 1, 1),
            fecha_fin=date(2025, 1, 1),
            estado='Finalizada',
            cupo_maximo=10,
        )

        persona_est = Persona.objects.create(dni='84000001', nombre='Ana', apellido='Est', correo='e@test.com')
        usuario_est = Usuario.objects.create(persona=persona_est, contrasena='pw')
        estudiante = Estudiante.objects.create(usuario=usuario_est, nivel_estudios='SE', institucion_actual='Colegio')
        inscripcion = Inscripcion.objects.create(estudiante=estudiante, comision=comision, estado='confirmado')
        Asistencia.objects.create(inscripcion=inscripcion, fecha_clase=date(2025, 1, 1), presente=True)
        registro, _ = RegistroAsistencia.objects.get_or_create(inscripcion=inscripcion)
        registro.calcular_porcentaje()

        response = self.client.get(reverse('administracion:exportar_estadisticas_estudiantes_curso'), secure=True)
        self.assertEqual(response.status_code, 200)
        self.assertIn('application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', response['Content-Type'])
        self.assertTrue(response.content.startswith(b'PK'))

        wb = load_workbook(filename=BytesIO(response.content))
        self.assertIn('Estudiantes por Curso', wb.sheetnames)
        self.assertIn('Detalle Estudiantes', wb.sheetnames)

        ws = wb['Estudiantes por Curso']
        self.assertEqual(ws.cell(row=1, column=1).value, 'Curso')
        self.assertEqual(ws.cell(row=2, column=1).value, curso.nombre)
        self.assertEqual(ws.cell(row=2, column=2).value, 1)
        self.assertEqual(ws.cell(row=2, column=3).value, 1)
